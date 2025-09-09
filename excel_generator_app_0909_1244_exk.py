# 代码生成时间: 2025-09-09 12:44:15
import os
from starlette.applications import Starlette
from starlette.requests import Request
from starlette.responses import FileResponse
from starlette.routing import Route
from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.writer.excel import save_workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

"""
Excel表格自动生成器，使用STARLETTE框架。
提供API接口以生成Excel文件。
"""

class ExcelGeneratorApp(Starlette):
    def __init__(self):
        super().__init__(
            routes=[
                Route('/', self.index),
                Route('/generate', self.generate_excel),
            ]
        )

    async def index(self, request: Request):
        """
        首页接口，返回欢迎信息。
        """
        return "Welcome to Excel Generator API"

    async def generate_excel(self, request: Request):
        """
        生成Excel文件的接口。
        """
        try:
            # 创建新的Excel工作簿
            wb = Workbook()
            # 选择默认工作表
            ws = wb.active
            # 添加标题行
            ws.title = 'Generated Excel'
            # 添加数据行
            for i in range(1, 11):
                ws[f'A{i}'] = f'Data {i}'
            # 保存工作簿
            temp_file_path = 'temp_excel.xlsx'
            save_workbook(wb, temp_file_path)
            # 返回文件响应
            return FileResponse(temp_file_path, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        except Exception as e:
            # 错误处理
            return {"error": str(e)}

    def run(self):
        """
        运行服务器。
        """
        if os.name != 'nt':
            import uvicorn
            uvicorn.run(self, host='0.0.0.0', port=8000)
        else:
            import http.server
            import socketserver
            with socketserver.TCPServer((